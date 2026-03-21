from flask import Blueprint, render_template, request, jsonify, redirect, url_for, flash, send_file
from app import db
from app.models import Cliente, Factura
from openpyxl import Workbook, load_workbook
from openpyxl.styles import PatternFill, Font, Alignment
from io import BytesIO
import os

bp = Blueprint('clientes', __name__)

@bp.route('/')
def listar():
    q = request.args.get('q', '')
    clientes = Cliente.query
    if q:
        clientes = clientes.filter(Cliente.nombre.contains(q) | 
                                  Cliente.ruc_ci.contains(q))
    clientes = clientes.order_by(Cliente.nombre).all()
    return render_template('clientes/list.html', clientes=clientes, q=q)

def _is_xhr():
    return request.headers.get('X-Requested-With') == 'XMLHttpRequest'

@bp.route('/nuevo', methods=['GET', 'POST'])
def nuevo():
    if request.method == 'POST':
        try:
            cliente = Cliente(
                nombre=request.form['nombre'].strip(),
                ruc_ci=request.form.get('ruc_ci', '').strip()
            )
            db.session.add(cliente)
            db.session.commit()
            if _is_xhr():
                return jsonify({'success': True, 'message': 'Cliente creado correctamente'})
            flash('Cliente creado correctamente', 'success')
            return redirect(url_for('clientes.listar'))
        except Exception as e:
            if _is_xhr():
                return jsonify({'success': False, 'message': str(e)}), 400
            raise
    return render_template('clientes/form.html')

@bp.route('/<int:id>/editar', methods=['GET', 'POST'])
def editar(id):
    cliente = Cliente.query.get_or_404(id)
    if request.method == 'POST':
        try:
            cliente.nombre = request.form['nombre'].strip()
            cliente.ruc_ci = request.form.get('ruc_ci', '').strip()
            db.session.commit()
            if _is_xhr():
                return jsonify({'success': True, 'message': 'Cliente actualizado correctamente'})
            flash('Cliente actualizado correctamente', 'success')
            return redirect(url_for('clientes.listar'))
        except Exception as e:
            if _is_xhr():
                return jsonify({'success': False, 'message': str(e)}), 400
            raise
    return render_template('clientes/form.html', cliente=cliente)

@bp.route('/<int:id>/eliminar', methods=['POST'])
def eliminar(id):
    cliente = Cliente.query.get_or_404(id)
    if cliente.facturas:
        flash('No se puede eliminar un cliente con facturas asociadas', 'error')
        return redirect(url_for('clientes.listar'))
    db.session.delete(cliente)
    db.session.commit()
    flash('Cliente eliminado correctamente', 'success')
    return redirect(url_for('clientes.listar'))

@bp.route('/<int:id>/historial')
def historial(id):
    cliente = Cliente.query.get_or_404(id)
    facturas = Factura.query.filter_by(cliente_id=id).order_by(Factura.fecha_emision.desc()).all()
    return render_template('clientes/historial.html', cliente=cliente, facturas=facturas)

@bp.route('/api/buscar')
def api_buscar():
    q = request.args.get('q', '')
    clientes = Cliente.query
    if q:
        clientes = clientes.filter(Cliente.nombre.contains(q))
    clientes = clientes.limit(10).all()
    return jsonify([c.to_dict() for c in clientes])

@bp.route('/importar', methods=['GET', 'POST'])
def importar():
    """Importar clientes desde Excel"""
    if request.method == 'POST':
        if 'archivo' not in request.files:
            msg = 'No se seleccionó ningún archivo'
            if _is_xhr():
                return jsonify({'success': False, 'message': msg}), 400
            flash(msg, 'error')
            return redirect(url_for('clientes.importar'))
        
        archivo = request.files['archivo']
        if archivo.filename == '':
            msg = 'No se seleccionó ningún archivo'
            if _is_xhr():
                return jsonify({'success': False, 'message': msg}), 400
            flash(msg, 'error')
            return redirect(url_for('clientes.importar'))
        
        try:
            wb = load_workbook(archivo, data_only=True)
            ws = wb.active
            
            # Buscar la fila de encabezados
            header_row = None
            for row_idx, row in enumerate(ws.iter_rows(min_row=1, max_row=20), 1):
                valores = [str(cell.value).lower() if cell.value else '' for cell in row]
                valores_normalizados = [
                    valor.replace('ó', 'o').replace('é', 'e').replace('í', 'i').replace(' ', '_')
                    for valor in valores if valor
                ]
                if 'nombre' in valores_normalizados and any(v in valores_normalizados for v in ('ci', 'cedula', 'ruc', 'ci/ruc', 'ci_ruc')):
                    header_row = row_idx
                    break
            
            if not header_row:
                msg = 'No se encontraron encabezados válidos. Use la plantilla con columnas: Nombre, CI'
                if _is_xhr():
                    return jsonify({'success': False, 'message': msg}), 400
                flash(msg, 'error')
                return redirect(url_for('clientes.importar'))
            
            # Leer encabezados (plantilla: solo Nombre y CI)
            def _norm(s):
                if s is None:
                    return ''
                return str(s).strip().lower().replace('ó', 'o').replace('é', 'e').replace('í', 'i')
            headers = [str(cell.value).strip() if cell.value else '' for cell in ws[header_row]]
            col_indices = {}
            for idx, header in enumerate(headers):
                if not header:
                    continue
                h = _norm(header)
                if h == 'nombre':
                    col_indices['nombre'] = idx
                elif h in ('ci', 'cedula', 'ruc'):
                    col_indices['ci'] = idx
            
            if 'nombre' not in col_indices:
                msg = 'Falta la columna "Nombre". Descargue la plantilla y no cambie los nombres de las columnas.'
                if _is_xhr():
                    return jsonify({'success': False, 'message': msg}), 400
                flash(msg, 'error')
                return redirect(url_for('clientes.importar'))
            
            # Procesar filas
            clientes_importados = 0
            clientes_actualizados = 0
            errores = []
            
            for row_idx, row in enumerate(ws.iter_rows(min_row=header_row + 1), header_row + 1):
                nombre = None
                ci = None
                try:
                    if 'nombre' in col_indices:
                        v = row[col_indices['nombre']].value
                        nombre = str(v).strip() if v is not None else None
                    if not nombre or str(nombre).lower() in ('', 'none', 'null'):
                        continue
                    if 'ci' in col_indices:
                        v = row[col_indices['ci']].value
                        ci = str(v).strip() if v is not None else None
                        if ci:
                            ci = ci.replace('-', '').replace(' ', '').strip()
                            if ci.lower() in ('none', 'null', ''):
                                ci = None
                    cliente_existente = None
                    if ci:
                        cliente_existente = Cliente.query.filter_by(ruc_ci=ci).first()
                    if not cliente_existente:
                        cliente_existente = Cliente.query.filter_by(nombre=nombre).first()
                    if cliente_existente:
                        cliente_existente.nombre = nombre
                        cliente_existente.ruc_ci = ci or ''
                        clientes_actualizados += 1
                    else:
                        nuevo_cliente = Cliente(nombre=nombre, ruc_ci=ci or '')
                        db.session.add(nuevo_cliente)
                        clientes_importados += 1
                        
                except Exception as e:
                    errores.append(f'Fila {row_idx}: {str(e)}')
                    continue
            
            db.session.commit()
            
            mensaje = f'Importación completada: {clientes_importados} nuevos, {clientes_actualizados} actualizados'
            if errores:
                mensaje += f'. Errores: {len(errores)}'
            if _is_xhr():
                return jsonify({'success': True, 'message': mensaje})
            flash(mensaje, 'success' if not errores else 'warning')
            return redirect(url_for('clientes.listar'))
            
        except Exception as e:
            db.session.rollback()
            if _is_xhr():
                return jsonify({'success': False, 'message': str(e)}), 400
            flash(f'Error al importar: {str(e)}', 'error')
            return redirect(url_for('clientes.importar'))
    
    return render_template('clientes/importar.html')

@bp.route('/descargar-plantilla')
def descargar_plantilla():
    """Genera y descarga la plantilla Excel para importar clientes"""
    # Crear workbook
    wb = Workbook()
    ws = wb.active
    ws.title = 'Plantilla Importación'
    
    # Estilos para encabezados
    header_fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
    header_font = Font(bold=True, color='FFFFFF', size=11)
    center_align = Alignment(horizontal='center', vertical='center')
    
    # Encabezados: solo Nombre y CI (deben coincidir exactamente con la importación)
    headers = ['Nombre', 'CI']
    ws.append(headers)
    
    # Aplicar estilos a encabezados
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center_align
    
    # Datos de ejemplo
    ejemplos = [
        ['Juan Pérez', '1234567'],
        ['María González', '7654321'],
        ['Carlos Rodríguez', '1122334'],
    ]
    
    for fila in ejemplos:
        ws.append(fila)
    
    # Ajustar ancho de columnas
    column_widths = [30, 20]
    for col_idx, width in enumerate(column_widths, 1):
        ws.column_dimensions[ws.cell(row=1, column=col_idx).column_letter].width = width
    
    # Guardar en memoria
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    
    return send_file(
        output,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name='plantilla_importacion_clientes.xlsx'
    )

