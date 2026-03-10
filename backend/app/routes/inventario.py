from flask import Blueprint, render_template, request, jsonify, redirect, url_for, flash, send_file
from app import db
from app.config import get_uploads_dir
from app.models import Producto
from sqlalchemy.exc import IntegrityError
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from werkzeug.utils import secure_filename
import os
import uuid

bp = Blueprint('inventario', __name__)

def _is_xhr():
    return request.headers.get('X-Requested-With') == 'XMLHttpRequest'

@bp.route('/')
def listar():
    q = request.args.get('q', '')
    productos = Producto.query.filter_by(activo=True)
    if q:
        productos = productos.filter(Producto.nombre.contains(q) | 
                                     Producto.codigo.contains(q))
    productos = productos.order_by(Producto.nombre).all()
    return render_template('inventario/list.html', productos=productos, q=q)

@bp.route('/nuevo', methods=['GET', 'POST'])
def nuevo():
    if request.method == 'POST':
        try:
            codigo = request.form['codigo'].strip()
            nombre = request.form['nombre'].strip()

            # Validar que el código no exista (constraint UNIQUE en la BD)
            existente = Producto.query.filter_by(codigo=codigo).first()
            if existente:
                msg = f'Ya existe un producto con el código "{codigo}".'
                if _is_xhr():
                    return jsonify({'success': False, 'message': msg}), 400
                flash(msg, 'error')
                return render_template('inventario/form.html')

            # Precio de compra (obligatorio)
            precio_compra = float(request.form['precio_compra'])
            # Precio de venta 1 (obligatorio)
            precio_1 = float(request.form['precio_1'])
            precio_2 = float(request.form.get('precio_2')) if request.form.get('precio_2') else None
            if precio_2 and precio_2 > precio_1:
                msg = 'El Precio 2 no puede ser mayor que el Precio 1.'
                if _is_xhr():
                    return jsonify({'success': False, 'message': msg}), 400
                flash(msg, 'error')
                return render_template('inventario/form.html')
            producto = Producto(
                codigo=codigo,
                nombre=nombre,
                # Guardamos explícitamente precio de compra; el alias precio_unitario sigue existiendo
                precio_compra=precio_compra,
                precio_1=precio_1,
                precio_2=precio_2,
                stock=int(request.form.get('stock', 0))
            )
            db.session.add(producto)
            try:
                db.session.commit()
            except IntegrityError:
                db.session.rollback()
                msg = f'Ya existe un producto con el código "{codigo}".'
                if _is_xhr():
                    return jsonify({'success': False, 'message': msg}), 400
                flash(msg, 'error')
                return render_template('inventario/form.html')
            if _is_xhr():
                return jsonify({'success': True, 'message': 'Producto creado correctamente'})
            flash('Producto creado correctamente', 'success')
            return redirect(url_for('inventario.listar'))
        except Exception as e:
            if _is_xhr():
                return jsonify({'success': False, 'message': str(e)}), 400
            raise
    return render_template('inventario/form.html')

@bp.route('/<int:id>/editar', methods=['GET', 'POST'])
def editar(id):
    producto = Producto.query.get_or_404(id)
    if request.method == 'POST':
        try:
            # Precio de compra (obligatorio)
            precio_compra = float(request.form['precio_compra'])
            # Precio de venta 1 (obligatorio)
            precio_1 = float(request.form['precio_1'])
            precio_2 = float(request.form.get('precio_2')) if request.form.get('precio_2') else None
            if precio_2 and precio_2 > precio_1:
                msg = 'El Precio 2 no puede ser mayor que el Precio 1.'
                if _is_xhr():
                    return jsonify({'success': False, 'message': msg}), 400
                flash(msg, 'error')
                return render_template('inventario/form.html', producto=producto)
            producto.nombre = request.form['nombre'].strip()
            producto.precio_compra = precio_compra
            producto.precio_1 = precio_1
            producto.precio_2 = precio_2
            producto.stock = int(request.form.get('stock', 0))
            db.session.commit()
            if _is_xhr():
                return jsonify({'success': True, 'message': 'Producto actualizado correctamente'})
            flash('Producto actualizado correctamente', 'success')
            return redirect(url_for('inventario.listar'))
        except Exception as e:
            if _is_xhr():
                return jsonify({'success': False, 'message': str(e)}), 400
            raise
    return render_template('inventario/form.html', producto=producto)

@bp.route('/<int:id>/eliminar', methods=['POST'])
def eliminar(id):
    producto = Producto.query.get_or_404(id)
    if producto.detalles:
        flash('No se puede eliminar un producto con facturas asociadas', 'error')
        return redirect(url_for('inventario.listar'))
    producto.activo = False
    db.session.commit()
    flash('Producto eliminado correctamente', 'success')
    return redirect(url_for('inventario.listar'))

@bp.route('/api/buscar')
def api_buscar():
    q = request.args.get('q', '')
    productos = Producto.query.filter_by(activo=True)
    if q:
        productos = productos.filter(Producto.nombre.contains(q) | 
                                    Producto.codigo.contains(q))
    productos = productos.limit(10).all()
    return jsonify([p.to_dict() for p in productos])

@bp.route('/api/<int:id>/stock')
def api_stock(id):
    producto = Producto.query.get_or_404(id)
    return jsonify({
        'producto_id': producto.id,
        'stock_disponible': producto.stock,
        'stock_bajo': producto.stock < 10
    })

@bp.route('/importar', methods=['GET', 'POST'])
def importar():
    if request.method == 'POST':
        if 'archivo' not in request.files:
            flash('No se seleccionó ningún archivo', 'error')
            return redirect(url_for('inventario.importar'))
        
        archivo = request.files['archivo']
        if archivo.filename == '':
            flash('No se seleccionó ningún archivo', 'error')
            return redirect(url_for('inventario.importar'))
        
        if not archivo.filename.endswith(('.xlsx', '.xls')):
            msg = 'El archivo debe ser Excel (.xlsx o .xls)'
            if _is_xhr():
                return jsonify({'success': False, 'message': msg}), 400
            flash(msg, 'error')
            return redirect(url_for('inventario.importar'))
        
        # Guardar archivo temporalmente
        filename = secure_filename(archivo.filename)
        temp_filename = f"{uuid.uuid4()}_{filename}"
        upload_folder = get_uploads_dir()
        filepath = os.path.join(upload_folder, temp_filename)
        archivo.save(filepath)
        
        try:
            # Leer archivo Excel
            wb = load_workbook(filepath, data_only=True)
            ws = wb.active
            
            # Buscar la fila de encabezados
            header_row = None
            for row_idx, row in enumerate(ws.iter_rows(min_row=1, max_row=20), 1):
                valores = [str(cell.value).lower() if cell.value else '' for cell in row]
                valores_normalizados = [
                    valor.replace('ó', 'o').replace('í', 'i').replace('é', 'e').replace('á', 'a').replace('ú', 'u').replace(' ', '_')
                    for valor in valores if valor
                ]
                if 'nombre' in valores_normalizados and ('codigo' in valores_normalizados or 'precio_compra' in valores_normalizados):
                    header_row = row_idx
                    break
            
            if not header_row:
                msg = 'No se encontraron encabezados. Use la plantilla con: Nombre, Código, Precio_compra, Precio_1, Precio_2, Stock'
                if _is_xhr():
                    return jsonify({'success': False, 'message': msg}), 400
                flash(msg, 'error')
                os.remove(filepath)
                return redirect(url_for('inventario.importar'))
            
            # Columnas según plantilla: Nombre, Código, Precio_compra, Precio_1, Precio_2, Stock (normalizar acentos y espacios)
            def norm(s):
                if s is None:
                    return ''
                return (str(s).strip().lower()
                        .replace('ó', 'o').replace('í', 'i').replace('é', 'e').replace('á', 'a').replace('ú', 'u')
                        .replace(' ', '_'))
            headers = [norm(cell.value) for cell in ws[header_row]]
            col_map = {
                'nombre': 'nombre',
                'codigo': 'codigo',
                'precio_compra': 'precio_compra',
                'precio_1': 'precio_1',
                'precio_2': 'precio_2',
                'stock': 'stock',
            }
            col_indices = {}
            for idx, h in enumerate(headers):
                if not h:
                    continue
                key = col_map.get(h, h)
                if key in col_map and key not in col_indices:
                    col_indices[key] = idx
            
            if 'nombre' not in col_indices:
                msg = 'Falta la columna "Nombre". Descargue la plantilla y no cambie los nombres de las columnas.'
                if _is_xhr():
                    return jsonify({'success': False, 'message': msg}), 400
                flash(msg, 'error')
                os.remove(filepath)
                return redirect(url_for('inventario.importar'))
            
            # Procesar filas
            productos_creados = 0
            productos_actualizados = 0
            errores = []
            
            for row_idx, row in enumerate(ws.iter_rows(min_row=header_row + 1), header_row + 1):
                # Obtener valores
                nombre = None
                codigo = None
                precio_compra = None
                precio_1 = None
                precio_2 = None
                stock = 0
                
                try:
                    if 'nombre' in col_indices:
                        nombre_cell = row[col_indices['nombre']]
                        nombre = str(nombre_cell.value).strip() if nombre_cell.value else None
                    
                    if not nombre or nombre.lower() in ['', 'none', 'null']:
                        continue
                    
                    if 'codigo' in col_indices:
                        codigo_cell = row[col_indices['codigo']]
                        codigo = str(codigo_cell.value).strip() if codigo_cell.value else None
                    
                    # Precio de compra (opcional en archivo, pero lo recomendamos)
                    if 'precio_compra' in col_indices:
                        precio_compra_cell = row[col_indices['precio_compra']]
                        try:
                            precio_compra = float(precio_compra_cell.value) if precio_compra_cell.value else None
                        except (ValueError, TypeError):
                            precio_compra = None

                    if 'precio_1' in col_indices:
                        precio_1_cell = row[col_indices['precio_1']]
                        try:
                            precio_1 = float(precio_1_cell.value) if precio_1_cell.value else None
                        except (ValueError, TypeError):
                            precio_1 = None
                    
                    # Leer precio_2 si existe la columna
                    if 'precio_2' in col_indices:
                        precio_2_cell = row[col_indices['precio_2']]
                        try:
                            precio_2 = float(precio_2_cell.value) if precio_2_cell.value else None
                        except (ValueError, TypeError):
                            precio_2 = None
                    
                    if 'stock' in col_indices:
                        stock_cell = row[col_indices['stock']]
                        try:
                            stock = int(float(stock_cell.value)) if stock_cell.value else 0
                        except (ValueError, TypeError):
                            stock = 0
                    
                    # Generar código si no existe
                    if not codigo:
                        codigo = f"PROD-{uuid.uuid4().hex[:8].upper()}"
                    
                    # Validar que P1 sea mayor o igual que P2
                    if precio_1 and precio_2 and precio_2 > precio_1:
                        errores.append(f"Fila {row_idx}: El Precio 2 no puede ser mayor que el Precio 1. P1 debe ser el precio más alto.")
                        continue
                    
                    # Asegurar que precio_1 tenga un valor (es obligatorio, precio de venta 1)
                    if not precio_1 or precio_1 == 0:
                        # Si no hay precio_1, intentar usar precio_compra como fallback
                        if precio_compra and precio_compra > 0:
                            precio_1 = precio_compra
                        else:
                            errores.append(f"Fila {row_idx}: El Precio venta 1 es obligatorio.")
                            continue

                    # Asegurar que precio_compra tenga valor (si falta, igualarlo a precio_1)
                    if not precio_compra or precio_compra == 0:
                        precio_compra = precio_1
                    
                    # Verificar si el producto ya existe (por código o nombre)
                    producto_existente = None
                    if codigo:
                        producto_existente = Producto.query.filter_by(codigo=codigo, activo=True).first()
                    
                    if not producto_existente:
                        producto_existente = Producto.query.filter_by(nombre=nombre, activo=True).first()
                    
                    if producto_existente:
                        # Actualizar producto existente
                        producto_existente.nombre = nombre
                        producto_existente.precio_compra = precio_compra
                        producto_existente.precio_1 = precio_1
                        producto_existente.precio_2 = precio_2
                        producto_existente.stock = stock
                        productos_actualizados += 1
                    else:
                        # Crear nuevo producto
                        nuevo_producto = Producto(
                            codigo=codigo,
                            nombre=nombre,
                            precio_compra=precio_compra,
                            precio_1=precio_1,
                            precio_2=precio_2 if precio_2 else None,
                            stock=stock
                        )
                        db.session.add(nuevo_producto)
                        productos_creados += 1
                
                except Exception as e:
                    errores.append(f"Fila {row_idx}: {str(e)}")
                    continue
            
            # Guardar cambios
            db.session.commit()
            
            # Eliminar archivo temporal
            os.remove(filepath)
            
            mensaje = f"Importación completada: {productos_creados} creados, {productos_actualizados} actualizados"
            if errores:
                mensaje += f". {len(errores)} errores."
            if _is_xhr():
                return jsonify({'success': True, 'message': mensaje})
            if errores:
                flash(mensaje, 'warning')
            else:
                flash(mensaje, 'success')
            return redirect(url_for('inventario.listar'))
        
        except Exception as e:
            try:
                if os.path.exists(filepath):
                    os.remove(filepath)
            except NameError:
                pass
            if _is_xhr():
                return jsonify({'success': False, 'message': str(e)}), 400
            flash(f'Error al procesar el archivo: {str(e)}', 'error')
            return redirect(url_for('inventario.importar'))
    
    return render_template('inventario/importar.html')

@bp.route('/descargar-plantilla')
def descargar_plantilla():
    """Plantilla Excel con columnas: Nombre, Código, Precio_compra, Precio_1, Precio_2, Stock"""
    wb = Workbook()
    ws = wb.active
    ws.title = 'Productos'
    header_fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
    header_font = Font(bold=True, color='FFFFFF', size=11)
    center_align = Alignment(horizontal='center', vertical='center')
    # Nombres exactos que el sistema lee (no cambiar)
    headers = ['Nombre', 'Código', 'Precio_compra', 'Precio_1', 'Precio_2', 'Stock']
    ws.append(headers)
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center_align
    ejemplos = [
        ['Producto 1', 'PROD-001', 50.00, 75.00, 65.00, 100],
        ['Producto 2', 'PROD-002', 30.00, 45.00, 40.00, 50],
    ]
    for fila in ejemplos:
        ws.append(fila)
    for col_idx, width in enumerate([25, 14, 14, 12, 12, 10], 1):
        ws.column_dimensions[ws.cell(row=1, column=col_idx).column_letter].width = width
    from io import BytesIO
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return send_file(
        output,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name='plantilla_productos.xlsx'
    )

@bp.route('/exportar')
def exportar():
    """Exporta el inventario actual a un archivo Excel"""
    q = request.args.get('q', '')
    
    productos_query = Producto.query.filter_by(activo=True)
    if q:
        productos_query = productos_query.filter(
            Producto.nombre.contains(q) | Producto.codigo.contains(q)
        )
    productos = productos_query.order_by(Producto.nombre).all()
    
    # Crear workbook en memoria
    wb = Workbook()
    ws = wb.active
    ws.title = 'Inventario'
    
    # Estilos de encabezado
    header_fill = PatternFill(start_color='2563EB', end_color='2563EB', fill_type='solid')
    header_font = Font(bold=True, color='FFFFFF', size=11)
    center_align = Alignment(horizontal='center', vertical='center')
    
    # Encabezados alineados con la plantilla de importación
    headers = [
        'Nombre',
        'Código',
        'Precio_compra',
        'Precio_1',
        'Precio_2',
        'Stock',
    ]
    ws.append(headers)
    
    for col_idx, _ in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center_align
    
    # Datos
    for producto in productos:
        # Precio de compra: usamos el campo dedicado si existe, o el alias precio_unitario
        precio_compra = getattr(producto, 'precio_compra', None)
        if not precio_compra:
            precio_compra = getattr(producto, 'precio_unitario', 0) or 0

        precio_1 = producto.precio_1 if producto.precio_1 is not None else ''
        precio_2 = producto.precio_2 if producto.precio_2 is not None else ''

        # Precio 3 es opcional: solo se llenará si el modelo/capa de datos lo define

        ws.append([
            producto.nombre,
            producto.codigo,
            precio_compra,
            precio_1,
            precio_2,
            producto.stock,
        ])
    
    # Ajustar ancho de columnas
    column_widths = [30, 18, 14, 14, 14, 10]
    for col_idx, width in enumerate(column_widths, 1):
        ws.column_dimensions[ws.cell(row=1, column=col_idx).column_letter].width = width
    
    # Guardar en memoria
    from io import BytesIO
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    
    return send_file(
        output,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name='inventario_sisfac.xlsx'
    )
