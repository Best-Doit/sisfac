from flask import Blueprint, render_template, request, flash, send_file, jsonify, redirect, url_for, make_response
from app import db
from app.models import Cliente, Producto, Factura, DetalleFactura, Talonario, Configuracion
from app.config import get_database_path, get_backups_dir
from sqlalchemy import text, func
import os
import shutil
import traceback
from datetime import datetime, date, timedelta
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, Alignment, PatternFill

bp = Blueprint('ajustes', __name__)


def _get_backups_list():
    """Lista de backups para usar en modales."""
    backups = []
    try:
        backups_dir = get_backups_dir()
        if os.path.exists(backups_dir):
            for file in os.listdir(backups_dir):
                if file.endswith('.db'):
                    file_path = os.path.join(backups_dir, file)
                    backups.append({
                        'nombre': file,
                        'fecha': datetime.fromtimestamp(os.path.getmtime(file_path)).strftime('%d/%m/%Y %H:%M:%S'),
                        'tamaño': os.path.getsize(file_path)
                    })
            backups.sort(key=lambda x: x['fecha'], reverse=True)
    except Exception:
        pass
    return backups


@bp.route('/')
def index():
    """Pantalla principal de ajustes: solo botones que abren modales."""
    return render_template('ajustes/index.html')


@bp.route('/modal/configuracion')
def modal_configuracion():
    """Fragmento HTML para modal de configuración (umbral stock)."""
    umbral_stock = Configuracion.obtener_int('umbral_stock_bajo', 10)
    return render_template('ajustes/modales/configuracion.html', umbral_stock=umbral_stock)


@bp.route('/modal/backup')
def modal_backup():
    """Fragmento HTML para modal de copias de seguridad."""
    backups = _get_backups_list()
    return render_template('ajustes/modales/backup.html', backups=backups)


@bp.route('/modal/restaurar')
def modal_restaurar():
    """Fragmento HTML para modal de restaurar backup."""
    return render_template('ajustes/modales/restaurar.html')


@bp.route('/modal/exportar')
def modal_exportar():
    """Fragmento HTML para modal de exportar a Excel."""
    return render_template('ajustes/modales/exportar.html')


@bp.route('/modal/importar')
def modal_importar():
    """Fragmento HTML para modal de importar datos desde Excel."""
    return render_template('ajustes/modales/importar.html')


@bp.route('/modal/borrar-datos')
def modal_borrar_datos():
    """Fragmento HTML para modal de zona peligrosa."""
    return render_template('ajustes/modales/borrar_datos.html')


@bp.route('/backup', methods=['POST'])
def crear_backup():
    """Crear una copia de seguridad de la base de datos"""
    try:
        db_path = get_database_path()
        if not os.path.exists(db_path):
            flash('No se encontró la base de datos', 'error')
            return jsonify({'success': False, 'message': 'No se encontró la base de datos'})
        
        backups_dir = get_backups_dir()
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        backup_filename = f'sisfac_backup_{timestamp}.db'
        backup_path = os.path.join(backups_dir, backup_filename)
        
        shutil.copy2(db_path, backup_path)
        
        flash(f'Backup creado: {backup_filename}', 'success')
        return jsonify({'success': True, 'message': f'Backup creado: {backup_filename}'})
    except Exception as e:
        flash(f'Error al crear backup: {str(e)}', 'error')
        return jsonify({'success': False, 'message': str(e)})

@bp.route('/restaurar', methods=['POST'])
def restaurar_backup():
    """Restaurar un backup"""
    try:
        if 'archivo' not in request.files:
            flash('No se seleccionó ningún archivo', 'error')
            return jsonify({'success': False, 'message': 'No se seleccionó ningún archivo'})
        
        archivo = request.files['archivo']
        if archivo.filename == '':
            flash('No se seleccionó ningún archivo', 'error')
            return jsonify({'success': False, 'message': 'No se seleccionó ningún archivo'})
        
        if not archivo.filename.endswith('.db'):
            flash('El archivo debe ser una base de datos (.db)', 'error')
            return jsonify({'success': False, 'message': 'El archivo debe ser una base de datos (.db)'})
        
        db_path = get_database_path()
        
        # Cerrar todas las conexiones de la base de datos antes de restaurar
        db.session.close()
        
        # Crear backup del estado actual antes de restaurar
        if os.path.exists(db_path):
            backups_dir = get_backups_dir()
            timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
            backup_actual = os.path.join(backups_dir, f'sisfac_antes_restaurar_{timestamp}.db')
            shutil.copy2(db_path, backup_actual)
        
        # Guardar el archivo subido
        archivo.save(db_path)
        
        flash('Backup restaurado correctamente. Reinicia la aplicación.', 'success')
        return jsonify({'success': True, 'message': 'Backup restaurado correctamente'})
    except Exception as e:
        flash(f'Error al restaurar backup: {str(e)}', 'error')
        return jsonify({'success': False, 'message': str(e)})

@bp.route('/descargar-backup/<nombre>')
def descargar_backup(nombre):
    """Descargar un archivo de backup"""
    try:
        backups_dir = get_backups_dir()
        backup_path = os.path.join(backups_dir, nombre)
        
        if not os.path.exists(backup_path) or not nombre.endswith('.db'):
            flash('Archivo no encontrado', 'error')
            return jsonify({'success': False, 'message': 'Archivo no encontrado'})
        
        return send_file(backup_path, as_attachment=True, download_name=nombre)
    except Exception as e:
        flash(f'Error al descargar backup: {str(e)}', 'error')
        return jsonify({'success': False, 'message': str(e)})

@bp.route('/borrar-datos', methods=['POST'])
def borrar_datos():
    """Borrar todos los datos de las tablas principales"""
    try:
        # Detectar si es una llamada AJAX/JSON o un POST normal
        is_ajax = request.is_json or request.headers.get('X-Requested-With') == 'XMLHttpRequest'

        # Obtener confirmación del formulario o JSON (acepta 'confirmar' o 'confirmacion')
        if request.is_json:
            data = request.get_json(silent=True) or {}
            confirmacion = data.get('confirmar') or data.get('confirmacion') or ''
        else:
            confirmacion = request.form.get('confirmar') or request.form.get('confirmacion') or ''
        
        print(f"🔍 Confirmación recibida: '{confirmacion}'")
        
        # Validar confirmación (acepta "borrar" o "BORRAR")
        confirmacion_limpia = confirmacion.strip().lower()
        if confirmacion_limpia != 'borrar':
            msg = f'Debes escribir "BORRAR" para confirmar. Recibido: "{confirmacion}"'
            print(f"❌ {msg}")
            if is_ajax:
                return jsonify({'success': False, 'message': msg})
            flash(msg, 'error')
            return redirect(url_for('ajustes.index'))
        
        print("✅ Confirmación válida, procediendo a borrar datos...")
        
        # Crear backup antes de borrar
        db_path = get_database_path()
        backup_creado = False
        if os.path.exists(db_path):
            try:
                backups_dir = get_backups_dir()
                timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
                backup_antes_borrar = os.path.join(backups_dir, f'sisfac_antes_borrar_{timestamp}.db')
                shutil.copy2(db_path, backup_antes_borrar)
                backup_creado = True
                print(f"✅ Backup creado: {backup_antes_borrar}")
            except Exception as e:
                print(f"⚠️ Error al crear backup: {e}")
                # Continuar aunque falle el backup
        
        # Deshabilitar temporalmente las foreign keys en SQLite
        try:
            db.session.execute(text('PRAGMA foreign_keys = OFF'))
            print("✅ Foreign keys deshabilitadas")
        except Exception as e:
            print(f"⚠️ Error al deshabilitar foreign keys: {e}")
        
        # Borrar datos usando SQL directo para evitar conflictos con
        # relaciones ORM (cascade, backref) en SQLAlchemy 2.x
        try:
            tablas = [
                ('detalle_factura', 'DetalleFactura'),
                ('facturas', 'Factura'),
                ('productos', 'Producto'),
                ('clientes', 'Cliente'),
                ('talonarios', 'Talonario'),
            ]
            for tabla, nombre in tablas:
                count = db.session.execute(text(f'SELECT COUNT(*) FROM {tabla}')).scalar()
                db.session.execute(text(f'DELETE FROM {tabla}'))
                print(f"🗑️ {nombre}: {count} eliminados")

            db.session.commit()
            print("✅ Cambios guardados correctamente")

        except Exception as e:
            db.session.rollback()
            print(f"❌ Error al borrar datos: {e}")
            traceback.print_exc()
            raise e
        finally:
            try:
                db.session.execute(text('PRAGMA foreign_keys = ON'))
            except Exception:
                pass
        
        msg_ok = 'Datos borrados correctamente.'
        if backup_creado:
            msg_ok += ' Se creó un backup automático.'
        
        print(f"✅ {msg_ok}")
        
        if is_ajax:
            return jsonify({'success': True, 'message': msg_ok})
        flash(msg_ok, 'success')
        return redirect(url_for('ajustes.index'))
    except Exception as e:
        db.session.rollback()
        msg_err = f'Error al borrar datos: {str(e)}'
        print(f"❌ {msg_err}")
        traceback.print_exc()
        if request.is_json or request.headers.get('X-Requested-With') == 'XMLHttpRequest':
            return jsonify({'success': False, 'message': msg_err})
        flash(msg_err, 'error')
        return redirect(url_for('ajustes.index'))

@bp.route('/configuracion', methods=['POST'])
def guardar_configuracion():
    """Guardar configuración del sistema"""
    try:
        umbral_stock = request.form.get('umbral_stock', '10')
        
        Configuracion.establecer('umbral_stock_bajo', umbral_stock, 'Umbral de stock bajo para alertas')
        
        flash('Configuración guardada correctamente', 'success')
        return jsonify({'success': True, 'message': 'Configuración guardada correctamente'})
    except Exception as e:
        flash(f'Error al guardar configuración: {str(e)}', 'error')
        return jsonify({'success': False, 'message': str(e)})

@bp.route('/exportar-datos')
def exportar_datos():
    """Exportar todos los datos a Excel"""
    try:
        wb = Workbook()
        
        # Hoja de Clientes (solo ID, Nombre y CI/RUC)
        ws_clientes = wb.active
        ws_clientes.title = "Clientes"
        headers = ['ID', 'Nombre', 'CI/RUC']
        ws_clientes.append(headers)
        for cliente in Cliente.query.all():
            ws_clientes.append([
                cliente.id,
                cliente.nombre,
                cliente.ruc_ci or ''
            ])
        
        # Hoja de Productos (sin columna Activo, pero manteniendo Fecha Registro para seguimiento)
        ws_productos = wb.create_sheet("Productos")
        headers = ['ID', 'Código', 'Nombre', 'Precio_compra', 'Precio_1', 'Precio_2', 'Stock', 'Activo', 'Fecha Registro']
        ws_productos.append(headers)
        for producto in Producto.query.all():
            ws_productos.append([
                producto.id,
                producto.codigo,
                producto.nombre,
                producto.precio_unitario,
                producto.precio_1 or '',
                producto.precio_2 or '',
                producto.stock,
                'Sí' if producto.activo else 'No',
                producto.fecha_registro.strftime('%d/%m/%Y %H:%M:%S') if producto.fecha_registro else ''
            ])
        
        # Hoja de Talonarios
        ws_talonarios = wb.create_sheet("Talonarios")
        headers = ['ID', 'Nombre', 'Prefijo', 'Número Inicio', 'Número Fin', 'Número Actual', 'Activo']
        ws_talonarios.append(headers)
        for talonario in Talonario.query.all():
            ws_talonarios.append([
                talonario.id,
                talonario.nombre,
                talonario.prefijo,
                talonario.numero_inicio,
                talonario.numero_fin,
                talonario.numero_actual,
                'Sí' if talonario.activo else 'No'
            ])

        # Hoja de Configuración
        ws_config = wb.create_sheet("Configuracion")
        headers = ['ID', 'Clave', 'Valor', 'Descripción', 'Fecha Actualización']
        ws_config.append(headers)
        for config in Configuracion.query.all():
            ws_config.append([
                config.id,
                config.clave,
                config.valor or '',
                config.descripcion or '',
                config.fecha_actualizacion.strftime('%d/%m/%Y %H:%M:%S') if config.fecha_actualizacion else ''
            ])
        
        # Hoja de Facturas
        ws_facturas = wb.create_sheet("Facturas")
        headers = ['ID', 'Número Factura', 'Cliente ID', 'Cliente Nombre', 'Talonario ID', 'Talonario Nombre', 
                  'Fecha Emisión', 'Fecha Vencimiento', 'Fecha Creación', 'Fecha Edición', 
                  'Subtotal', 'IVA', 'Total', 'Estado', 'Notas']
        ws_facturas.append(headers)
        for factura in Factura.query.all():
            ws_facturas.append([
                factura.id,
                factura.numero_factura,
                factura.cliente_id,
                factura.cliente.nombre if factura.cliente else '',
                factura.talonario_id or '',
                factura.talonario.nombre if factura.talonario else '',
                factura.fecha_emision.strftime('%d/%m/%Y') if factura.fecha_emision else '',
                factura.fecha_vencimiento.strftime('%d/%m/%Y') if factura.fecha_vencimiento else '',
                factura.fecha_creacion.strftime('%d/%m/%Y %H:%M:%S') if factura.fecha_creacion else '',
                factura.fecha_edicion.strftime('%d/%m/%Y %H:%M:%S') if factura.fecha_edicion else '',
                factura.subtotal,
                factura.iva,
                factura.total,
                factura.estado,
                factura.notas or ''
            ])
        
        # Hoja de Detalles de Factura
        ws_detalles = wb.create_sheet("DetallesFactura")
        headers = ['ID', 'Factura ID', 'Número Factura', 'Producto ID', 'Producto Código', 'Producto Nombre', 'Cantidad', 'Precio Unitario', 'Subtotal']
        ws_detalles.append(headers)
        for detalle in DetalleFactura.query.all():
            ws_detalles.append([
                detalle.id,
                detalle.factura_id,
                detalle.factura.numero_factura if detalle.factura else '',
                detalle.producto_id,
                detalle.producto.codigo if detalle.producto else '',
                detalle.producto.nombre if detalle.producto else '',
                detalle.cantidad,
                detalle.precio_unitario,
                detalle.subtotal
            ])
        
        # Estilizar encabezados
        for ws in wb.worksheets:
            header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            header_font = Font(bold=True, color="FFFFFF")
            for cell in ws[1]:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal="center")
        
        # Guardar en memoria
        from io import BytesIO
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        
        fecha = datetime.now().strftime('%Y%m%d_%H%M%S')
        filename = f'sisfac_datos_export_{fecha}.xlsx'
        
        response = make_response(output.getvalue())
        response.headers['Content-Type'] = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        response.headers['Content-Disposition'] = f'attachment; filename={filename}'
        
        flash('Datos exportados correctamente', 'success')
        return response
    except Exception as e:
        flash(f'Error al exportar datos: {str(e)}', 'error')
        return redirect(url_for('ajustes.index'))


@bp.route('/importar-datos', methods=['POST'])
def importar_datos():
    """Importar todos los datos desde un Excel exportado por el sistema"""
    try:
        if 'archivo' not in request.files:
            flash('No se seleccionó ningún archivo', 'error')
            return redirect(url_for('ajustes.index'))
        
        archivo = request.files['archivo']
        if archivo.filename == '':
            flash('No se seleccionó ningún archivo', 'error')
            return redirect(url_for('ajustes.index'))
        
        if not archivo.filename.endswith(('.xlsx', '.xls')):
            flash('El archivo debe ser Excel (.xlsx o .xls)', 'error')
            return redirect(url_for('ajustes.index'))
        
        wb = load_workbook(archivo, data_only=True)
        required_sheets = ['Clientes', 'Productos', 'Talonarios', 'Facturas', 'DetallesFactura']
        faltantes = [s for s in required_sheets if s not in wb.sheetnames]
        if faltantes:
            flash(f'Faltan hojas requeridas en el Excel: {", ".join(faltantes)}', 'error')
            return redirect(url_for('ajustes.index'))
        
        def _norm(value):
            if value is None:
                return ''
            s = str(value).strip().lower()
            s = (s.replace('á', 'a').replace('é', 'e').replace('í', 'i')
                  .replace('ó', 'o').replace('ú', 'u').replace('ñ', 'n'))
            s = s.replace('/', '_').replace(' ', '_')
            return s
        
        def _parse_int(value, default=None):
            if value in (None, ''):
                return default
            try:
                return int(float(value))
            except Exception:
                return default
        
        def _parse_float(value, default=None):
            if value in (None, ''):
                return default
            try:
                return float(value)
            except Exception:
                return default
        
        def _parse_bool(value, default=True):
            if value is None or value == '':
                return default
            if isinstance(value, bool):
                return value
            if isinstance(value, (int, float)):
                return value != 0
            s = str(value).strip().lower()
            if s in ('si', 'sí', 'true', '1', 'yes', 'y'):
                return True
            if s in ('no', 'false', '0', 'n'):
                return False
            return default
        
        def _parse_date(value):
            if value in (None, ''):
                return None
            if isinstance(value, datetime):
                return value.date()
            if isinstance(value, date):
                return value
            s = str(value).strip()
            for fmt in ('%d/%m/%Y', '%Y-%m-%d', '%d-%m-%Y'):
                try:
                    return datetime.strptime(s, fmt).date()
                except Exception:
                    pass
            return None
        
        def _parse_datetime(value):
            if value in (None, ''):
                return None
            if isinstance(value, datetime):
                return value
            if isinstance(value, date):
                return datetime.combine(value, datetime.min.time())
            s = str(value).strip()
            for fmt in ('%d/%m/%Y %H:%M:%S', '%Y-%m-%d %H:%M:%S', '%d/%m/%Y', '%Y-%m-%d'):
                try:
                    return datetime.strptime(s, fmt)
                except Exception:
                    pass
            return None
        
        def _leer_filas(sheet, mapping):
            headers = [_norm(c.value) for c in sheet[1]]
            idx_map = {h: i for i, h in enumerate(headers) if h}
            filas = []
            for row in sheet.iter_rows(min_row=2):
                data = {}
                empty = True
                for col_key, field in mapping.items():
                    idx = idx_map.get(col_key)
                    if idx is None:
                        continue
                    value = row[idx].value
                    if value not in (None, ''):
                        empty = False
                    if field:
                        data[field] = value
                if not empty:
                    filas.append(data)
            return filas
        
        clientes_map = {
            'id': 'id',
            'nombre': 'nombre',
            'ci_ruc': 'ruc_ci',
            'ci': 'ruc_ci',
            'ruc': 'ruc_ci',
        }
        productos_map = {
            'id': 'id',
            'codigo': 'codigo',
            'nombre': 'nombre',
            'precio_compra': 'precio_compra',
            'precio_principal': 'precio_principal',
            'precio_unitario': 'precio_principal',
            'precio_1': 'precio_1',
            'precio_p1': 'precio_1',
            'precio_2': 'precio_2',
            'precio_p2': 'precio_2',
            'stock': 'stock',
            'activo': 'activo',
            'fecha_registro': 'fecha_registro'
        }
        talonarios_map = {
            'id': 'id',
            'nombre': 'nombre',
            'prefijo': 'prefijo',
            'numero_inicio': 'numero_inicio',
            'numero_fin': 'numero_fin',
            'numero_actual': 'numero_actual',
            'activo': 'activo'
        }
        facturas_map = {
            'id': 'id',
            'numero_factura': 'numero_factura',
            'cliente_id': 'cliente_id',
            'talonario_id': 'talonario_id',
            'fecha_emision': 'fecha_emision',
            'fecha_vencimiento': 'fecha_vencimiento',
            'fecha_creacion': 'fecha_creacion',
            'fecha_edicion': 'fecha_edicion',
            'subtotal': 'subtotal',
            'iva': 'iva',
            'total': 'total',
            'estado': 'estado',
            'notas': 'notas'
        }
        detalles_map = {
            'id': 'id',
            'factura_id': 'factura_id',
            'producto_id': 'producto_id',
            'cantidad': 'cantidad',
            'precio_unitario': 'precio_unitario',
            'subtotal': 'subtotal'
        }
        config_map = {
            'id': 'id',
            'clave': 'clave',
            'valor': 'valor',
            'descripcion': 'descripcion',
            'fecha_actualizacion': 'fecha_actualizacion'
        }
        
        errores = []
        
        # Deshabilitar foreign keys y limpiar tablas
        db.session.execute(text('PRAGMA foreign_keys = OFF'))
        db.session.query(DetalleFactura).delete(synchronize_session=False)
        db.session.query(Factura).delete(synchronize_session=False)
        db.session.query(Producto).delete(synchronize_session=False)
        db.session.query(Cliente).delete(synchronize_session=False)
        db.session.query(Talonario).delete(synchronize_session=False)
        db.session.query(Configuracion).delete(synchronize_session=False)
        
        # Importar Clientes (solo ID, Nombre y CI/RUC)
        for row in _leer_filas(wb['Clientes'], clientes_map):
            nombre = str(row.get('nombre') or '').strip()
            if not nombre:
                continue
            cliente = Cliente(
                nombre=nombre,
                ruc_ci=str(row.get('ruc_ci') or '').strip(),
            )
            row_id = _parse_int(row.get('id'))
            if row_id is not None:
                cliente.id = row_id
            db.session.add(cliente)
        
        # Importar Productos
        for row in _leer_filas(wb['Productos'], productos_map):
            codigo = str(row.get('codigo') or '').strip()
            nombre = str(row.get('nombre') or '').strip()
            if not codigo or not nombre:
                errores.append('Productos: falta Código o Nombre en una fila')
                continue
            precio_compra = _parse_float(row.get('precio_compra'))
            precio_principal = _parse_float(row.get('precio_principal'))
            precio_1 = _parse_float(row.get('precio_1'))
            precio_2 = _parse_float(row.get('precio_2'))
            precio_unitario = (
                precio_compra
                if precio_compra is not None else
                (precio_principal if precio_principal is not None else (precio_1 if precio_1 is not None else 0.0))
            )
            producto = Producto(
                codigo=codigo,
                nombre=nombre,
                precio_compra=precio_unitario,
                precio_1=precio_1,
                precio_2=precio_2,
                stock=_parse_int(row.get('stock'), 0)
            )
            row_id = _parse_int(row.get('id'))
            if row_id is not None:
                producto.id = row_id
            fecha_registro = _parse_datetime(row.get('fecha_registro'))
            if fecha_registro:
                producto.fecha_registro = fecha_registro
            producto.activo = _parse_bool(row.get('activo'), True)
            db.session.add(producto)
        
        # Importar Talonarios
        for row in _leer_filas(wb['Talonarios'], talonarios_map):
            nombre = str(row.get('nombre') or '').strip()
            prefijo = str(row.get('prefijo') or '').strip()
            if not nombre or not prefijo:
                errores.append('Talonarios: falta Nombre o Prefijo en una fila')
                continue
            numero_inicio = _parse_int(row.get('numero_inicio'), 1)
            numero_fin = _parse_int(row.get('numero_fin'), numero_inicio)
            numero_actual = _parse_int(row.get('numero_actual'), numero_inicio)
            talonario = Talonario(
                nombre=nombre,
                prefijo=prefijo,
                numero_inicio=numero_inicio,
                numero_fin=numero_fin,
                numero_actual=numero_actual,
                activo=_parse_bool(row.get('activo'), True)
            )
            row_id = _parse_int(row.get('id'))
            if row_id is not None:
                talonario.id = row_id
            db.session.add(talonario)
        
        # Importar Facturas
        for row in _leer_filas(wb['Facturas'], facturas_map):
            numero_factura = str(row.get('numero_factura') or '').strip()
            cliente_id = _parse_int(row.get('cliente_id'))
            fecha_emision = _parse_date(row.get('fecha_emision'))
            if not numero_factura or cliente_id is None or fecha_emision is None:
                errores.append('Facturas: falta Número, Cliente o Fecha Emisión en una fila')
                continue
            factura = Factura(
                numero_factura=numero_factura,
                cliente_id=cliente_id,
                talonario_id=_parse_int(row.get('talonario_id')),
                fecha_emision=fecha_emision,
                fecha_vencimiento=_parse_date(row.get('fecha_vencimiento')),
                subtotal=_parse_float(row.get('subtotal'), 0.0) or 0.0,
                iva=_parse_float(row.get('iva'), 0.0) or 0.0,
                total=_parse_float(row.get('total'), 0.0) or 0.0,
                estado=str(row.get('estado') or 'PAGADA'),
                notas=str(row.get('notas') or '')
            )
            row_id = _parse_int(row.get('id'))
            if row_id is not None:
                factura.id = row_id
            fecha_creacion = _parse_datetime(row.get('fecha_creacion'))
            fecha_edicion = _parse_datetime(row.get('fecha_edicion'))
            if fecha_creacion:
                factura.fecha_creacion = fecha_creacion
            if fecha_edicion:
                factura.fecha_edicion = fecha_edicion
            db.session.add(factura)
        
        # Importar Detalles de Factura
        for row in _leer_filas(wb['DetallesFactura'], detalles_map):
            factura_id = _parse_int(row.get('factura_id'))
            producto_id = _parse_int(row.get('producto_id'))
            if factura_id is None or producto_id is None:
                errores.append('DetallesFactura: falta Factura ID o Producto ID en una fila')
                continue
            cantidad = _parse_int(row.get('cantidad'), 0) or 0
            precio_unitario = _parse_float(row.get('precio_unitario'), 0.0) or 0.0
            subtotal = _parse_float(row.get('subtotal'), None)
            if subtotal is None:
                subtotal = cantidad * precio_unitario
            detalle = DetalleFactura(
                factura_id=factura_id,
                producto_id=producto_id,
                cantidad=cantidad,
                precio_unitario=precio_unitario,
                subtotal=subtotal
            )
            row_id = _parse_int(row.get('id'))
            if row_id is not None:
                detalle.id = row_id
            db.session.add(detalle)
        
        # Importar Configuración (opcional)
        if 'Configuracion' in wb.sheetnames:
            for row in _leer_filas(wb['Configuracion'], config_map):
                clave = str(row.get('clave') or '').strip()
                if not clave:
                    continue
                config = Configuracion(
                    clave=clave,
                    valor=str(row.get('valor') or ''),
                    descripcion=str(row.get('descripcion') or '')
                )
                row_id = _parse_int(row.get('id'))
                if row_id is not None:
                    config.id = row_id
                fecha_actualizacion = _parse_datetime(row.get('fecha_actualizacion'))
                if fecha_actualizacion:
                    config.fecha_actualizacion = fecha_actualizacion
                db.session.add(config)
        
        if errores:
            db.session.rollback()
            msg = f'Error en la importación: {errores[0]}'
            flash(msg, 'error')
            return redirect(url_for('ajustes.index'))
        
        db.session.commit()
        flash('Datos importados correctamente. Reinicia la aplicación para aplicar cambios.', 'success')
        return redirect(url_for('ajustes.index'))
    except Exception as e:
        db.session.rollback()
        flash(f'Error al importar datos: {str(e)}', 'error')
        return redirect(url_for('ajustes.index'))
    finally:
        try:
            db.session.execute(text('PRAGMA foreign_keys = ON'))
        except Exception:
            pass


@bp.route('/limpiar-backups', methods=['POST'])
def limpiar_backups():
    """Eliminar backups antiguos (más de 30 días)"""
    try:
        backups_dir = get_backups_dir()
        if not os.path.exists(backups_dir):
            return jsonify({'success': False, 'message': 'Directorio de backups no encontrado'})
        
        fecha_limite = datetime.now() - timedelta(days=30)
        eliminados = 0
        
        for file in os.listdir(backups_dir):
            if file.endswith('.db'):
                file_path = os.path.join(backups_dir, file)
                fecha_modificacion = datetime.fromtimestamp(os.path.getmtime(file_path))
                if fecha_modificacion < fecha_limite:
                    os.remove(file_path)
                    eliminados += 1
        
        flash(f'Se eliminaron {eliminados} backup(s) antiguo(s)', 'success')
        return jsonify({'success': True, 'message': f'Se eliminaron {eliminados} backup(s) antiguo(s)'})
    except Exception as e:
        flash(f'Error al limpiar backups: {str(e)}', 'error')
        return jsonify({'success': False, 'message': str(e)})
